#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl import load_workbook

# load workbook
wbFile = "./testsheet01.xlsx"
wb = load_workbook(wbFile)

# scan sheets
print(str(wb.sheetnames))

# set sheet active and get title
sheet1 = wb.active
sheet1.title

# read from sheet
number = 2
while number <= 11:
    print('Date: ' + str(sheet1["A" + str(number)].value))
    print(str(sheet1["B1"].value) + ':' + str(sheet1["B" + str(number)].value))
    print(str(sheet1["C1"].value) + ':' + str(sheet1["C" + str(number)].value))
    print(str(sheet1["D1"].value) + ':' + str(sheet1["D" + str(number)].value))
    number += 1

# calculate sums and write to sheet
